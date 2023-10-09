from flask import Flask, render_template, request, send_file
import pdfreader
import openpyxl
import io
import re
import pdfplumber


app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    manufacturer = request.form['manufacturer']
    pdf_file = request.files['pdf_file']

    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.cell(row=1, column=1).value = "Manufacturer"
    worksheet.cell(row=1, column=2).value = manufacturer

    with pdfreader.PdfReader(pdf_file) as pdf:
        full_text = ''
        for page in pdf.pages:
            full_text += page.extract_text()


    workbook.save(output)

    output.seek(0)
    return send_file(output, attachment_filename='results.xlsx', as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
